import openpyxl
import pytz
from collections import OrderedDict
import dateutil.parser
from simple_salesforce import Salesforce

# instantiate simple_salesforce object with appropriate Salesforce login credentials
sf = Salesforce(
    username='integration@levelsolar.com',
    password='HrNt7DrqaEfZmBqJRan9dKFzmQFp',
    security_token='yWlJG8lAKCq1pTBkbBMSVcKg')

# get data from Salesforce using instantiated object
employees = sf.query_all("SELECT Name, Office__c, Role__c, CAD_Specialist__c, CountyRegion__c, CountyRegion__r.Name FROM Employee__c WHERE Status__c = 'Active'")["records"]
metrics = sf.query_all("SELECT Date__c, Ambassador__r.Name, Office__c, AmbShifts__c, Shift_Length__c, Doors__c, Appointments__c FROM Metrics__c WHERE Ambassador__c != null AND AmbShifts__c = 1 AND Office__c != null AND Date__c < TODAY AND (Date__c = THIS_MONTH OR Date__c = THIS_WEEK)")["records"]
leads = sf.query_all("SELECT Lead_Number__c, Ambassador__r.Name, Ambassador__r.Office__c, Sit_Date__c FROM Lead WHERE (Sit_Date__c = THIS_MONTH OR Sit_Date__c = THIS_WEEK)")["records"]
sales = sf.query_all("SELECT Account_Number__c, CreatedDate, Town_Permit_Submitted__c, InstallDate__c, Ambassador__r.Name, SalesRepE__r.Name, Ambassador__r.Office__c, ZipCodeRegion__r.Name, SalesRepE__r.CountyRegion__r.Name FROM Opportunity WHERE (CreatedDate = THIS_WEEK or CreatedDate = THIS_MONTH or Town_Permit_Submitted__c = THIS_WEEK or Town_Permit_Submitted__c = THIS_MONTH or InstallDate__c = THIS_WEEK or InstallDate__c = THIS_MONTH) AND Is_Additional_Meter__c = null AND CreatedDate < TODAY")["records"]
appointments = sf.query_all("SELECT Lead__r.Lead_Number__c, Lead__r.Ambassador__r.Office__c, Lead__r.Ambassador__r.Name, Lead__r.Sit_Date__c, Lead__r.ZipCodeRegion__r.Name, ScheduledDate__c, Outcome__c, Outcome_Submitted__r.Name, Outcome_Submitted__r.CountyRegion__c, Assigned_To__r.Name, Assigned_To__r.CountyRegion__r.Name, Assigned_To__r.Office__c, Confirmed__c, Canceled__c, Rescheduled__c FROM Interaction__c WHERE Subject__c = 'Closer Appointment' AND Lead__c != null AND Lead__r.ZipCodeRegion__r.Name != null AND (ScheduledDate__c = THIS_WEEK OR ScheduledDate__c = THIS_MONTH OR Lead__r.Sit_Date__c = THIS_WEEK or Lead__r.Sit_Date__c = THIS_MONTH) AND ScheduledDate__c < TODAY")["records"]
cads = sf.query_all("SELECT Opportunity__r.Account_Number__c, Opportunity__r.Ambassador__r.Name, Opportunity__r.Ambassador__r.Office__c, Opportunity__r.ZipCodeRegion__r.Name, Opportunity__r.SalesRepE__r.Name, Outcome_Submitted__r.Name, Outcome_Submitted__r.CountyRegion__c, InteractionDate__c, Outcome__c, Confirmed__c, Canceled__c, Rescheduled__c FROM Interaction__c WHERE Subject__c = 'CAD Appointment' AND Opportunity__c != null AND (InteractionDate__c = THIS_WEEK OR InteractionDate__c = THIS_MONTH) and InteractionDate__c < TODAY")["records"]


# open pre-existing Excel workbook with openpyxl
wb = openpyxl.load_workbook('Sales Report.xlsx')

# list data sheets to be refreshed
data_sheets = ['Employees - Raw', 'Knocking Metrics - Raw', 'Leads - Raw', 'Sales Appointments - Raw', 'CAD Appointments - Raw', 'Opportunities - Raw']

# define time zones, for treatment of datetimes returned from Salesforce in UTC
utc_zone = pytz.timezone('UTC')
est_zone = pytz.timezone('US/Eastern')

# clears sheets, passed in as a list
def clear_sheets(titles):
	for title in titles:
		wb.remove_sheet(wb.get_sheet_by_name(title))
		wb.create_sheet(title=title)

# clear all sheets to be refreshed
clear_sheets(data_sheets)

# fills a single cell value; called by fill_sheet
def fill_cell(data, sheet, row, meta_dict):
	col_index = 1
	for key in meta_dict:
		if "." not in meta_dict[key]:
			v = data[row][meta_dict[key]]
			sheet.cell(row=row+2, column=col_index).value = v
			col_index += 1
		else:
			keys = meta_dict[key].split('.')
			if data[row][keys[0]]:
				if str(type(data[row][keys[0]][keys[1]])) != "<class 'collections.OrderedDict'>":
					v = data[row][keys[0]][keys[1]]
				elif str(type(data[row][keys[0]][keys[1]][keys[2]])) != "<class 'collections.OrderedDict'>":
					v = data[row][keys[0]][keys[1]][keys[2]]
				else:
					v = data[row][keys[0]][keys[1]][keys[2]][keys[3]]
			else:
				v = None
			sheet.cell(row=row+2, column=col_index).value = v
			col_index += 1

# fills in sheet with data from Salesforce; takes as arguments (a) simple_salesforce response, (b) sheet name, (c) ordered dictionary that relates column titles to Salesforce fields
def fill_sheet(data, sheet_name, meta_dict):
	sheet = wb.get_sheet_by_name(str(sheet_name))
	col = 1
	# fill in headers
	for key in meta_dict:
		cell = openpyxl.utils.get_column_letter(col) + '1'
		sheet[cell] = key
		col += 1
	# fill in data
	total_rows = len(data)
	for row in range(0, total_rows):
		fill_cell(data, sheet, row, meta_dict)

# refreshes columns in summary tabs with employee records
def fill_column(data, sheet_name, field_column_map, row, criteria_field=None):
	sheet = wb.get_sheet_by_name(str(sheet_name))
	for key in field_column_map:
		column_row = row
		field = key
		column = field_column_map[key]
		for i in range(row, 1001):
			sheet.cell(row=i, column=column).value = None
			i += 1
		if criteria_field != None:
			for record in data:
				if record[criteria_field] != None:
					if "." in field:
						field_keys = field.split('.')
						if record[field_keys[0]]:
							v = record[field_keys[0]][field_keys[1]]
							sheet.cell(row=column_row, column=column).value = v
							column_row += 1
					else:
						v = record[field]
						sheet.cell(row=column_row, column=column).value = v
						column_row += 1
		else:
			for record in data:
				if "." in field:
					field_keys = field.split('.')
					if record[field_keys[0]]:
						v = record[field_keys[0]][field_keys[1]]
						sheet.cell(row=column_row, column=column).value = v
						column_row += 1
				else:
					v = record[field]
					sheet.cell(row=column_row, column=column).value = v
					column_row += 1

# turns Salesforce datetime response into Python/Excel readable datetime
def parse_datetime(datetime):
	try:
		dt = dateutil.parser.parse(datetime)
		if dt.tzinfo != None:
			dt = dt.astimezone(est_zone).replace(tzinfo=None).replace(hour=0, minute=0, second=0, microsecond=0)
		return dt
	except:
		return None

# turns datetime fields into Python/Excel readable datetimes
def format_datetimes(data, date_fields):
	for record in data:
		for date_field in date_fields:
			if "." not in date_field:
				v = record[date_field]
				record[date_field] = parse_datetime(v)
			else:
				keys = date_field.split('.')
				if len(keys) == 2:
					v = record[keys[0]][keys[1]]
					record[keys[0]][keys[1]] = parse_datetime(v)
				elif len(keys) == 3:
					v = record[keys[0]][keys[1]][keys[2]]
					record[keys[0]][keys[1]][keys[2]] = parse_datetime(v)

# conbines 'Western Suffolk' and 'Eastern Suffolk' into 'Suffolk'
def clean_regions(data, region_fields):
	for region_field in region_fields:
		if '.' not in region_field:
			for record in data:
				if record[region_field] in ['Eastern Suffolk', 'Western Suffolk']:
					record[region_field] = 'Suffolk'
		else:
			keys = region_field.split('.')
			if len(keys) == 2:
				for record in data:
					try:
						if record[keys[0]][keys[1]] in ['Eastern Suffolk', 'Western Suffolk']:
							record[keys[0]][keys[1]] = 'Suffolk'
					except:
						pass
			elif len(keys) == 3:
				for record in data:
					try:
						if record[keys[0]][keys[1]][keys[2]] in ['Eastern Suffolk', 'Western Suffolk']:
							record[keys[0]][keys[1]][keys[2]] = 'Suffolk'
					except:
						pass

def de_dupe(data, field):
	ids = []
	if "." not in field:
		for record in data:
			ids.append(record[field])
		for record in data:
			count = ids.count(record[field])
			weight = round(1/float(count), 2)
			record['Weight'] = weight
	else:
		keys = field.split(".")
		for record in data:
			ids.append(record[keys[0]][keys[1]])
		for record in data:
			count = ids.count(record[keys[0]][keys[1]])
			weight = round(1/float(count), 2)
			record['Weight'] = weight

def metric_eligibility(data):
	sit_outcomes = ["Sale", "Sat, No Verification", "Failed Credit", "Passed Credit, No Sale", "Follow-Up Set", "Permitting / CO Issues"]
	for record in data:
		if record['Outcome__c'] in sit_outcomes:
			record["Appointment Sit"] = 1
		else:
			record["Appointment Sit"] = 0
		if record["Outcome__c"] in ["Sale", "Failed Credit", "Passed Credit, No Sale"]:
			record["Run Credit"] = 1
		else:
			record["Run Credit"] = 0
		if record["Outcome__c"] == "Failed Credit":
			record["Failed Credit"] = 1
		else:
			record["Failed Credit"] = 0
		if record["Outcome_Submitted__r"]:
			if record["Outcome_Submitted__r"]["CountyRegion__c"] != None and record["Outcome__c"] not in ["Postponed", "No Show"]:
				record["CAD Sit"] = 1
			else:
				record["CAD Sit"] = 0
		else:
			record["CAD Sit"] = 0

# prep and fill in 'Knocking Data - Raw' sheet
format_datetimes(metrics, ['Date__c'])
fill_sheet(metrics, 'Knocking Metrics - Raw', OrderedDict([('Date', 'Date__c'), ('Name', 'Ambassador__r.Name'), ('Office', 'Office__c'), ('Shifts', 'AmbShifts__c'), ('Shift Length', 'Shift_Length__c'), ('Doors', 'Doors__c'), ('Sets', 'Appointments__c')]))

# prep and fill in 'Sales Data - Raw' sheet
format_datetimes(sales, ['CreatedDate', 'Town_Permit_Submitted__c', 'InstallDate__c'])
clean_regions(sales, ['ZipCodeRegion__r.Name', 'SalesRepE__r.CountyRegion__r.Name'])
fill_sheet(sales, 'Opportunities - Raw', OrderedDict([('Account Number', 'Account_Number__c'), ('Sale Date', 'CreatedDate'), ('Ambassador', 'Ambassador__r.Name'), ('Ambassador - Office', 'Ambassador__r.Office__c'), ('Market', 'ZipCodeRegion__r.Name'), ('Consultant', 'SalesRepE__r.Name'), ('Consultant - Office', 'SalesRepE__r.CountyRegion__r.Name'), ('Permit Submitted', 'Town_Permit_Submitted__c'), ('Install Date', 'InstallDate__c')]))

# prep and fill in 'Employees - Raw' sheet
clean_regions(employees, ['CountyRegion__r.Name'])
fill_sheet(employees, 'Employees - Raw', OrderedDict([('Name', 'Name'), ('Role', 'Role__c'), ('Ambassador - Office', 'Office__c'), ('Consultant - Office', 'CountyRegion__r.Name'), ('CAD Specialist?', 'CAD_Specialist__c')]))

# prep and fill in 'Leads - Raw' sheet
format_datetimes(leads, ['Sit_Date__c'])
fill_sheet(leads, 'Leads - Raw', OrderedDict([('Lead Number', 'Lead_Number__c'), ('Ambassador', 'Ambassador__r.Name'), ('Office', 'Ambassador__r.Office__c'), ('Sit Date', 'Sit_Date__c')]))

# prep and fill in 'Appointment Data' - Raw' sheet
format_datetimes(appointments, ['ScheduledDate__c', 'Lead__r.Sit_Date__c'])
de_dupe(appointments, 'Lead__r.Lead_Number__c')
clean_regions(appointments, ['Lead__r.ZipCodeRegion__r.Name', 'Assigned_To__r.CountyRegion__r.Name'])
metric_eligibility(appointments)
fill_sheet(appointments, 'Sales Appointments - Raw', OrderedDict([('Lead Number', 'Lead__r.Lead_Number__c'), ('Ambassador', 'Lead__r.Ambassador__r.Name'), ('Office', 'Lead__r.Ambassador__r.Office__c'), ('Market', 'Lead__r.ZipCodeRegion__r.Name'), ('Scheduled Date', 'ScheduledDate__c'), ('Assigned To', 'Assigned_To__r.Name'), ('Assigned To - Office', 'Assigned_To__r.CountyRegion__r.Name'), ('Outcome', 'Outcome__c'), ('Consultant', 'Outcome_Submitted__r.Name'), ('Confirmed', 'Confirmed__c'), ('Canceled', 'Canceled__c'), ('Rescheduled', 'Rescheduled__c'), ('Lead Sit Date', 'Lead__r.Sit_Date__c'), ('De-Dupe Weight', 'Weight'), ('Appointment Sit', 'Appointment Sit'), ('Failed Credit', 'Failed Credit'), ('Run Credit', 'Run Credit')]))

# prep and fill in 'CAD Data - Raw' sheet
format_datetimes(cads, ['InteractionDate__c'])
clean_regions(cads, ['Opportunity__r.ZipCodeRegion__r.Name'])
metric_eligibility(cads)
fill_sheet(cads, 'CAD Appointments - Raw', OrderedDict([('Account Number', 'Opportunity__r.Account_Number__c'), ('Ambassador', 'Opportunity__r.Ambassador__r.Name'), ('Ambassador Office', 'Opportunity__r.Ambassador__r.Office__c'), ('Market', 'Opportunity__r.ZipCodeRegion__r.Name'), ('Consultant', 'Opportunity__r.SalesRepE__r.Name'), ('CAD Specialist', 'Outcome_Submitted__r.Name'), ('Outcome Date', 'InteractionDate__c'), ('Outcome', 'Outcome__c'), ('Confirmed', 'Confirmed__c'), ('Canceled', 'Canceled__c'), ('Rescheduled', 'Rescheduled__c'), ('CAD Sit', 'CAD Sit')]))

# fill ambassadors in 'Ambassadors - By Rep' sheet
fill_column(employees,'Ambassadors - By Rep', {'Name': 1, 'Office__c': 2}, 10, 'Office__c')

# fill consultants in 'Ambassadors - By Rep' sheet
fill_column(employees,'Consultants - By Rep', {'Name': 1, 'CountyRegion__r.Name': 2}, 10, 'CountyRegion__c')

# save Excel workbook
wb.save('Sales Report.xlsx')
