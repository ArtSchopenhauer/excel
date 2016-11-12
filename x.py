import openpyxl
import pytz
from collections import OrderedDict
import dateutil.parser
from simple_salesforce import Salesforce

# instantiate simple_salesforce object with appropriate Salesforce login credentials
sf = Salesforce(
    username='integration@levelsolar.com',
    password='HrNt7DrqaEfZmBqJRan9dKFzmQFp',
    security_token='yWlJG8lAKCq1pTBkbBMSVcKg')

# get data from Salesforce using instantiated object
ambassadors = sf.query_all("SELECT Name, Office__c, Role__c FROM Employee__c WHERE Status__c = 'Active' AND Office__c != null")["records"]
metrics = sf.query_all("SELECT Date__c, Ambassador__r.Name, Office__c, AmbShifts__c, Shift_Length__c, Doors__c, Appointments__c FROM Metrics__c WHERE Ambassador__c != null AND AmbShifts__c = 1 AND Office__c != null AND (Date__c = THIS_MONTH OR Date__c = THIS_WEEK)")["records"]
sales = sf.query_all("SELECT Account_Number__c, CreatedDate, Town_Permit_Submitted__c, Ambassador__r.Name, SalesRepE__r.Name, ZipCodeRegion__r.Name FROM Opportunity WHERE (CreatedDate = THIS_WEEK or CreatedDate = THIS_MONTH or Town_Permit_Submitted__c = THIS_WEEK or Town_Permit_Submitted__c = THIS_MONTH) AND ZipCodeRegion__c != null")["records"]
appointments = sf.query_all("SELECT Lead__r.Lead_Number__c, Lead__r.ZipCodeRegion__r.Name, Lead__r.Ambassador__r.Name, Lead__r.Sit_Date__c, ScheduledDate__c, Outcome__c, Outcome_Submitted__r.Name, Confirmed__c, Canceled__c, Rescheduled__c, Assigned_To__c FROM Interaction__c WHERE Subject__c = 'Closer Appointment' AND Lead__r.ZipCodeRegion__c != null AND (ScheduledDate__c = THIS_WEEK OR ScheduledDate__c = THIS_MONTH)")["records"]

# open pre-existing Excel workbook with openpyxl
wb = openpyxl.load_workbook('Sales Report.xlsx')

# list data sheets to be refreshed
data_sheets = ['Knocking Data - Raw', 'Sales Data - Raw', 'Ambassadors - Raw', 'Appointment Data - Raw']

# define time zones, for treatment of datetimes returned from Salesforce in UTC
utc_zone = pytz.timezone('UTC')
est_zone = pytz.timezone('US/Eastern')

# clears sheets, passed in as a list
def clear_sheets(titles):
	for title in titles:
		wb.remove_sheet(wb.get_sheet_by_name(title))
		wb.create_sheet(title=title)

# clear all sheets to be refreshed
clear_sheets(data_sheets)

# fills a single cell value; called by fill_sheet
def fill_cell(data, sheet, row, meta_dict):
	col_index = 1
	for key in meta_dict:
		if "." not in meta_dict[key]:
			v = data[row][meta_dict[key]]
			sheet.cell(row=row+2, column=col_index).value = v
			col_index += 1
		else:
			keys = meta_dict[key].split('.')
			if data[row][keys[0]]:
				if str(type(data[row][keys[0]][keys[1]])) != "<class 'collections.OrderedDict'>":
					v = data[row][keys[0]][keys[1]]
				else:
					v = data[row][keys[0]][keys[1]][keys[2]]
			else:
				v = None
			sheet.cell(row=row+2, column=col_index).value = v
			col_index += 1

# fills in sheet with data from Salesforce; takes as arguments (a) simple_salesforce response, (b) sheet name, (c) ordered dictionary that relates column titles to Salesforce fields
def fill_sheet(data, sheet_name, meta_dict):
	sheet = wb.get_sheet_by_name(str(sheet_name))
	col = 1
	# fill in headers
	for key in meta_dict:
		cell = openpyxl.utils.get_column_letter(col) + '1'
		sheet[cell] = key
		col += 1
	# fill in data
	total_rows = len(data)
	for row in range(0, total_rows):
		fill_cell(data, sheet, row, meta_dict)

# turns Saleseforce field into Python/Excel datetime
def format_datetimes(data, date_fields):
	for record in data:
		for date_field in date_fields:
			try:
				date = dateutil.parser.parse(record[date_field])
				if date.tzinfo != None:
					date = date.astimezone(est_zone).replace(tzinfo=None)
					date = date.replace(hour=0, minute=0, second=0, microsecond=0)
				record[date_field] = date
			except:
				pass

# conbines 'Western Suffolk' and 'Eastern Suffolk' into 'Suffolk'
def clean_regions(data, region_field):
	if '.' not in region_field:
		for record in data:
			if record[region_field] in ['Eastern Suffolk', 'Western Suffolk']:
				record[region_field] = 'Suffolk'
	else:
		keys = region_field.split('.')
		if len(keys) == 2:
			for record in data:
				if record[keys[0]][keys[1]] in ['Eastern Suffolk', 'Western Suffolk']:
					record[keys[0]][keys[1]] = 'Suffolk'
		elif len(keys) == 3:
			for record in data:
				if record[keys[0]][keys[1]][keys[2]] in ['Eastern Suffolk', 'Western Suffolk']:
					record[keys[0]][keys[1]][keys[2]] = 'Suffolk'

# prep and fill in 'Knocking Data - Raw' sheet
format_datetimes(metrics, ['Date__c'])
fill_sheet(metrics, 'Knocking Data - Raw', OrderedDict([('Date', 'Date__c'), ('Name', 'Ambassador__r.Name'), ('Office', 'Office__c'), ('Shifts', 'AmbShifts__c'), ('Shift Length', 'Shift_Length__c'), ('Doors', 'Doors__c'), ('Sets', 'Appointments__c')]))

# prep and fill in 'Sales Data - Raw' sheet
format_datetimes(sales, ['CreatedDate', 'Town_Permit_Submitted__c'])
clean_regions(sales, 'ZipCodeRegion__r.Name')
fill_sheet(sales, 'Sales Data - Raw', OrderedDict([('Account Number', 'Account_Number__c'), ('Sale Date', 'CreatedDate'), ('Ambassador', 'Ambassador__r.Name'), ('Market', 'ZipCodeRegion__r.Name'), ('Consultant', 'SalesRepE__r.Name'), ('Permit Submitted', 'Town_Permit_Submitted__c')]))

# prep and fill in 'Ambassadors - Raw' sheet
fill_sheet(ambassadors, 'Ambassadors - Raw', OrderedDict([('Name', 'Name'), ('Role', 'Role__c'), ('Office', 'Office__c')]))

# prep and fill in 'Appointment Data' - Raw' sheet
clean_regions(appointments, 'Lead__r.ZipCodeRegion__r.Name')
fill_sheet(appointments, 'Appointment Data - Raw', OrderedDict([('Lead Number', 'Lead__r.Lead_Number__c'), ('Market', 'Lead__r.ZipCodeRegion__r.Name')]))

# save Excel workbook
wb.save('Sales Report.xlsx')